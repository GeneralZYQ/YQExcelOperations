import openpyxl
import csv


def get_colums(workBook, sheetName):
    workSheet = workBook[sheetName]
    iterRows = workSheet.iter_rows(min_row=1, max_row=1)
    first_row = next(iterRows)
    columns = [cell.value for cell in first_row]
    return {sheetName: columns}


def fetchAllSheetNames(fileName):
    workBook = openpyxl.load_workbook(fileName)
    return workBook.sheetnames

def fetchAllRowsIn(filename, sheetname):
    workBook = openpyxl.load_workbook(filename)
    workSheet = workBook[sheetname]

    lines = []

    for row in workSheet.iter_rows():
        line = []
        for cell in row:
            line.append(cell.value)

        lines.append(line)

    return lines




def fetchAllColumnNamesInSheet(fileName, sheetName):
    workBook = openpyxl.load_workbook(fileName)
    columnNames = []
    if sheetName is None:

        for name in names:
            columns = get_colums(workBook, name)
            columnNames.append({name: columns})

        return columnNames

    else:
        columns = get_colums(workBook, sheetName)
        return columns

def vlookupWith(originalFilename, originalSheetName, originalColumnName, destinationFilename, destinationSheetName, destinationBenchmark, destinationColumnName):
    originalColumnNames = fetchAllColumnNamesInSheet(originalFilename, originalSheetName)[originalSheetName]
    originalIndex = originalColumnNames.index(originalColumnName)
    originalRows = fetchAllRowsIn(originalFilename, originalSheetName)

    destinationColumnNames = fetchAllColumnNamesInSheet(destinationFilename, destinationSheetName)[destinationSheetName]
    destinationBenchIndex = destinationColumnNames.index(destinationBenchmark)
    destinationCIndex = destinationColumnNames.index(destinationColumnName)
    destinationRows = fetchAllRowsIn(destinationFilename, destinationSheetName)

    for originalRow in originalRows:
        for destinationRow in destinationRows:
            if originalRow[originalIndex] == destinationRow[destinationBenchIndex]:
                originalRow.append(destinationRow[destinationCIndex])

        if len(originalRow) <= len(originalColumnNames):
            originalRow.append('Not Found')

    return originalRows



